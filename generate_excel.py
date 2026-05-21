#!/usr/bin/env python3
"""
Generates an interactive Excel dashboard from the latest ISCC certificates CSV.

Dashboard sheet layout:
  Left  — Select Country  → Pie chart: certificates by CB for that country
  Right — Select CB       → Bar chart: certificates by Country for that CB
"""

import csv
import glob
from collections import defaultdict
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Style constants ─────────────────────────────────────────────────────────
BLUE        = "005798"
LIGHT_BLUE  = "DCE6F1"
DARK_TEXT   = "1F3864"
WHITE       = "FFFFFF"
STRIPE      = "EBF3FB"
TOP_N       = 10   # CBs / countries shown per chart

# ── Helpers ──────────────────────────────────────────────────────────────────
def hdr(cell, bg=BLUE, fg=WHITE):
    cell.font      = Font(color=fg, bold=True, name="Calibri", size=11)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def label(cell, text, bold=False, size=11, color=DARK_TEXT):
    cell.value = text
    cell.font  = Font(bold=bold, name="Calibri", size=size, color=color)

def number(cell, val, size=14, color=BLUE):
    cell.value = val
    cell.font  = Font(bold=True, name="Calibri", size=size, color=color)

def find_csv():
    dated = sorted(glob.glob("ISCC certificates 20*.csv"))
    return dated[-1] if dated else "ISCC certificates latest.csv"


# ── Load & aggregate ─────────────────────────────────────────────────────────
def load_data(path):
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            country = (r.get("Country") or "Unknown").strip() or "Unknown"
            cb      = (r.get("Issuing CB") or "Unknown").strip() or "Unknown"
            rows.append({"country": country, "cb": cb})
    return rows


def aggregate(rows):
    """Returns:
       country_totals  — [(country, count), ...] sorted by count DESC
       cb_totals       — [(cb, count), ...] sorted by count DESC
       cb_by_country   — {country: [(cb, count), ...]} top-N sorted DESC (padded to TOP_N)
       country_by_cb   — {cb: [(country, count), ...]} top-N sorted DESC (padded to TOP_N)
    """
    country_cb_count = defaultdict(lambda: defaultdict(int))
    for r in rows:
        country_cb_count[r["country"]][r["cb"]] += 1

    country_totals = sorted(
        [(c, sum(v.values())) for c, v in country_cb_count.items()],
        key=lambda x: -x[1],
    )
    cb_totals_dict = defaultdict(int)
    for row in rows:
        cb_totals_dict[row["cb"]] += 1
    cb_totals = sorted(cb_totals_dict.items(), key=lambda x: -x[1])

    # Top-N CBs per country, padded
    cb_by_country = {}
    for country, _ in country_totals:
        pairs = sorted(country_cb_count[country].items(), key=lambda x: -x[1])[:TOP_N]
        # pad
        while len(pairs) < TOP_N:
            pairs.append(("", 0))
        cb_by_country[country] = pairs

    # Top-N countries per CB
    cb_country_count = defaultdict(lambda: defaultdict(int))
    for r in rows:
        cb_country_count[r["cb"]][r["country"]] += 1
    country_by_cb = {}
    for cb, _ in cb_totals:
        pairs = sorted(cb_country_count[cb].items(), key=lambda x: -x[1])[:TOP_N]
        while len(pairs) < TOP_N:
            pairs.append(("", 0))
        country_by_cb[cb] = pairs

    return country_totals, cb_totals, cb_by_country, country_by_cb


# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(rows, country_totals, cb_totals, cb_by_country, country_by_cb, csv_path,
                title="ISCC Active Certificates — Interactive Dashboard",
                dated_out=None, latest_out=None):
    wb = Workbook()

    # ── Hidden helper: CB_Data (country → top CBs) ──
    ws_cbd = wb.create_sheet("CB_Data")
    ws_cbd.append(["Country", "CB", "Count"])
    for country, _ in country_totals:
        for cb, cnt in cb_by_country[country]:
            ws_cbd.append([country, cb, cnt])
    ws_cbd.sheet_state = "hidden"

    # ── Hidden helper: Country_Data (CB → top countries) ──
    ws_cyd = wb.create_sheet("Country_Data")
    ws_cyd.append(["CB", "Country", "Count"])
    for cb, _ in cb_totals:
        for country, cnt in country_by_cb[cb]:
            ws_cyd.append([cb, country, cnt])
    ws_cyd.sheet_state = "hidden"

    # ── Hidden helper: Country_List ──
    ws_cl = wb.create_sheet("Country_List")
    for i, (c, _) in enumerate(country_totals, start=1):
        ws_cl.cell(row=i, column=1).value = c
    ws_cl.sheet_state = "hidden"

    # ── Hidden helper: CB_List ──
    ws_cbl = wb.create_sheet("CB_List")
    for i, (cb, _) in enumerate(cb_totals, start=1):
        ws_cbl.cell(row=i, column=1).value = cb
    ws_cbl.sheet_state = "hidden"

    # ── Country Summary sheet ──
    ws_cs = wb.create_sheet("Country Summary")
    ws_cs.append(["Country", "Active Certificates"])
    hdr(ws_cs["A1"]); hdr(ws_cs["B1"])
    ws_cs.column_dimensions["A"].width = 30
    ws_cs.column_dimensions["B"].width = 20
    for i, (c, cnt) in enumerate(country_totals, start=2):
        ws_cs.cell(row=i, column=1).value = c
        ws_cs.cell(row=i, column=2).value = cnt
        if i % 2 == 0:
            for col in [1, 2]:
                ws_cs.cell(row=i, column=col).fill = PatternFill("solid", fgColor=STRIPE)

    # ── Data sheet ──
    ws_data = wb.create_sheet("Data")
    fieldnames = ["Client Name", "Scope", "Issuing CB", "Expiry Date", "Country"]
    ws_data.append(fieldnames)
    for col_idx, h in enumerate(fieldnames, 1):
        hdr(ws_data.cell(row=1, column=col_idx))
    with open(csv_path, newline="", encoding="utf-8") as f:
        for row_idx, r in enumerate(csv.DictReader(f), start=2):
            for col_idx, key in enumerate(fieldnames, 1):
                ws_data.cell(row=row_idx, column=col_idx).value = r.get(key, "")
    n_data_rows = row_idx
    col_widths = [50, 25, 45, 15, 20]
    for i, w in enumerate(col_widths, 1):
        ws_data.column_dimensions[get_column_letter(i)].width = w
    tab = Table(displayName="CertData", ref=f"A1:{get_column_letter(len(fieldnames))}{n_data_rows}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True,
        showFirstColumn=False, showLastColumn=False,
    )
    ws_data.add_table(tab)

    # ── Dashboard sheet (first) ──
    ws = wb.active
    ws.title = "Dashboard"

    # Push Dashboard to front
    wb.move_sheet("Dashboard", offset=-(len(wb.sheetnames) - 1))

    # Row heights / column widths
    ws.row_dimensions[1].height = 44
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[7].height = 22
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 45
    ws.column_dimensions["G"].width = 5
    ws.column_dimensions["H"].width = 14

    # ── Title ──
    ws.merge_cells("A1:H1")
    ws["A1"].value     = title
    ws["A1"].font      = Font(bold=True, size=18, color=WHITE, name="Calibri")
    ws["A1"].fill      = PatternFill("solid", fgColor=BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # ── KPI row ──
    label(ws["B3"], "Total Active Certificates", bold=True, size=10, color="888888")
    number(ws["C3"], len(rows))
    label(ws["E3"], "Countries", bold=True, size=10, color="888888")
    number(ws["F3"], len(country_totals))

    # ── LEFT: Country selector ──
    label(ws["B5"], "Select Country:", bold=True, size=12)
    ws["C5"].value = country_totals[0][0]           # default = most-certified country
    ws["C5"].font  = Font(bold=True, name="Calibri", size=12)
    ws["C5"].fill  = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["C5"].alignment = Alignment(horizontal="left", vertical="center")

    dv_country = DataValidation(
        type="list",
        formula1=f"Country_List!$A$1:$A${len(country_totals)}",
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(dv_country)
    dv_country.sqref = "C5"

    # ── Country total ──
    label(ws["B6"], "Total certificates:", bold=False, size=10, color="888888")
    ws["C6"].value     = "=COUNTIF(CertData[Country],$C$5)"
    ws["C6"].font      = Font(bold=True, name="Calibri", size=12, color=BLUE)
    ws["C6"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[6].height = 20

    # CB breakdown table
    hdr(ws["B7"], bg=BLUE); ws["B7"].value = "Certification Body"
    hdr(ws["C7"], bg=BLUE); ws["C7"].value = "Count"

    for i in range(TOP_N):
        row = 8 + i
        off = i
        # CB name
        f_cb = f'=IFERROR(INDEX(CB_Data!$B:$B,MATCH($C$5,CB_Data!$A:$A,0)+{off}),"")'
        # Count — use NA() so pie chart skips empty slices
        f_cnt = (
            f"=IFERROR("
            f"IF(INDEX(CB_Data!$C:$C,MATCH($C$5,CB_Data!$A:$A,0)+{off})=0,"
            f"NA(),"
            f"INDEX(CB_Data!$C:$C,MATCH($C$5,CB_Data!$A:$A,0)+{off})),"
            f"NA())"
        )
        ws.cell(row=row, column=2).value = f_cb
        ws.cell(row=row, column=3).value = f_cnt
        bg = STRIPE if i % 2 == 0 else WHITE
        for col in [2, 3]:
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=bg)
            ws.cell(row=row, column=col).font = Font(name="Calibri", size=10)

    # Pie chart (CB distribution for selected country)
    pie = PieChart()
    pie.title  = "Certificates by CB (selected country)"
    pie.style  = 10
    pie.width  = 18
    pie.height = 14
    pie_data   = Reference(ws, min_col=3, min_row=7, max_row=7 + TOP_N)
    pie_labels = Reference(ws, min_col=2, min_row=8, max_row=7 + TOP_N)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_labels)
    # openpyxl defaults to numRef for categories; Excel requires strRef for text labels
    for s in pie.series:
        s.cat = AxDataSource(strRef=StrRef(f="'Dashboard'!$B$8:$B$17"))
    ws.add_chart(pie, "B19")

    # ── RIGHT: CB selector ──
    label(ws["E5"], "Select CB:", bold=True, size=12)
    ws["F5"].value = cb_totals[0][0]                # default = largest CB
    ws["F5"].font  = Font(bold=True, name="Calibri", size=12)
    ws["F5"].fill  = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["F5"].alignment = Alignment(horizontal="left", vertical="center")

    dv_cb = DataValidation(
        type="list",
        formula1=f"CB_List!$A$1:$A${len(cb_totals)}",
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(dv_cb)
    dv_cb.sqref = "F5"

    # Country breakdown table
    hdr(ws["E7"], bg=BLUE); ws["E7"].value = "Country"
    hdr(ws["F7"], bg=BLUE); ws["F7"].value = "Count"
    hdr(ws["G7"], bg=BLUE); ws["G7"].value = ""  # spacer

    for i in range(TOP_N):
        row = 8 + i
        off = i
        f_cy = f'=IFERROR(INDEX(Country_Data!$B:$B,MATCH($F$5,Country_Data!$A:$A,0)+{off}),"")'
        f_cnt = (
            f"=IFERROR("
            f"IF(INDEX(Country_Data!$C:$C,MATCH($F$5,Country_Data!$A:$A,0)+{off})=0,"
            f"NA(),"
            f"INDEX(Country_Data!$C:$C,MATCH($F$5,Country_Data!$A:$A,0)+{off})),"
            f"NA())"
        )
        ws.cell(row=row, column=5).value = f_cy
        ws.cell(row=row, column=6).value = f_cnt
        bg = STRIPE if i % 2 == 0 else WHITE
        for col in [5, 6]:
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=bg)
            ws.cell(row=row, column=col).font = Font(name="Calibri", size=10)

    # Bar chart (countries for selected CB)
    bar = BarChart()
    bar.type      = "bar"
    bar.grouping  = "clustered"
    bar.title     = "Top Countries for Selected CB"
    bar.style     = 10
    bar.width     = 18
    bar.height    = 14
    bar_data      = Reference(ws, min_col=6, min_row=7, max_row=7 + TOP_N)
    bar_labels    = Reference(ws, min_col=5, min_row=8, max_row=7 + TOP_N)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_labels)
    # Fix category axis to use strRef (text labels) instead of numRef
    for s in bar.series:
        s.cat = AxDataSource(strRef=StrRef(f="'Dashboard'!$E$8:$E$17"))
    ws.add_chart(bar, "E19")

    # ── Save ──
    if dated_out is None:
        date_part = csv_path.replace("ISCC certificates ", "").replace(".csv", "")
        dated_out = f"ISCC certificates {date_part}.xlsx"
    if latest_out is None:
        latest_out = "ISCC certificates latest.xlsx"
    for path in (dated_out, latest_out):
        wb.save(path)
        print(f"Saved → {path}")


def main():
    csv_path = find_csv()
    print(f"Reading {csv_path} …")
    rows = load_data(csv_path)
    print(f"Loaded {len(rows)} records. Aggregating …")
    country_totals, cb_totals, cb_by_country, country_by_cb = aggregate(rows)
    print(f"  {len(country_totals)} countries  |  {len(cb_totals)} CBs")
    build_excel(rows, country_totals, cb_totals, cb_by_country, country_by_cb, csv_path)


if __name__ == "__main__":
    main()
