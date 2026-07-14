#!/usr/bin/env python3
"""
Rebuilds the per-scheme certificate CSVs from the committed `<Scheme>
certificates latest.xlsx` dashboards — each dashboard's "Data" sheet holds every
row exactly as the scraper wrote it — and bundles them into a single dated zip.

Used by the monthly email-digest workflow so it can attach all CSVs without
re-running the (expensive) scrapers: the scrape workflows already commit fresh
dashboards each month, and this just repackages them.
"""

import csv
import os
import sys
import zipfile
from datetime import datetime, timezone

from openpyxl import load_workbook

SCHEMES = ["ISCC", "SURE", "PEFC", "FSC", "GGL", "SBP"]
EXPORT_DIR = "csv_export"


def xlsx_data_to_csv(xlsx_path, csv_path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["Data"]
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            rows = 0
            for row in ws.iter_rows(values_only=True):
                if row is None or all(c is None or c == "" for c in row):
                    continue                        # skip trailing blank rows
                w.writerow(["" if c is None else c for c in row])
                rows += 1
        return rows - 1                             # minus the header row
    finally:
        wb.close()


def main():
    os.makedirs(EXPORT_DIR, exist_ok=True)
    produced = []
    for scheme in SCHEMES:
        xlsx = f"{scheme} certificates latest.xlsx"
        if not os.path.exists(xlsx):
            print(f"  skip {scheme}: {xlsx} not found")
            continue
        csv_path = os.path.join(EXPORT_DIR, f"{scheme} certificates latest.csv")
        n = xlsx_data_to_csv(xlsx, csv_path)
        produced.append(csv_path)
        print(f"  {scheme}: {n} rows → {csv_path}")

    if not produced:
        sys.exit("No dashboards found to export — did the scrapers run?")

    date_str = datetime.now(timezone.utc).strftime("%Y.%m.%d")
    zip_path = f"certificate-csvs-{date_str}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for p in produced:
            z.write(p, arcname=os.path.basename(p))
    size_mb = os.path.getsize(zip_path) / 1e6
    print(f"Wrote {zip_path} — {len(produced)} CSVs, {size_mb:.1f} MB")


if __name__ == "__main__":
    main()
