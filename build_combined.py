#!/usr/bin/env python3
"""
Builds a single combined workbook from every scheme's committed
`<Scheme> certificates latest.xlsx` dashboard.

Two files are produced (both in one openpyxl session, so the dashboard charts are
never lost to a reload):

  • "All certificates latest.xlsx" — the full workbook:
      - Dashboard    interactive front page (Select Country → pie of certificates
                     by scheme; Select Scheme → bar of top countries), same design
                     as the per-scheme dashboards but driven by the combined set.
      - Data         one normalised, filterable row per certificate across all
                     schemes (Scheme, Identifier, Name, Country, Type,
                     Certification Body, Status, Valid From, Valid To).
      - Certification Bodies
                     filterable table of each certification body → record count
                     → which schemes report it (ISCC, SURE, GGL, SBP publish a
                     CB; PEFC and FSC do not).
      - ISCC … SBP   full native columns per scheme, so no detail is lost.
      - Summary      record counts per scheme.
  • "All certificates (dashboard) latest.xlsx" — the same but without the six
    per-scheme detail sheets, so it's small enough (~10 MB) to email.

The combined set's second dashboard *selector* dimension is Scheme (Certification
Body is absent for FSC/PEFC — 89% of rows — so an interactive per-country CB
split would be mostly empty). The Dashboard still surfaces CBs via a static
"Top Certification Bodies" table + bar, and the "Certification Bodies" sheet has
the full breakdown.
"""

import csv
import glob
import os
import sys
from datetime import date, datetime, timezone

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from generate_excel import aggregate, build_excel, hdr

SCHEMES = ["ISCC", "SURE", "PEFC", "FSC", "GGL", "SBP"]
BLUE, WHITE, STRIPE = "005798", "FFFFFF", "EBF3FB"

COMMON = ["Scheme", "Identifier", "Name", "Country", "Type",
          "Certification Body", "Status", "Valid From", "Valid To"]
COMMON_WIDTHS = (10, 20, 42, 18, 26, 26, 12, 13, 13)
COUNTRY_I, SCHEME_I = COMMON.index("Country"), COMMON.index("Scheme")
CB_I = COMMON.index("Certification Body")
DATE_COLS = {"Valid From", "Valid To"}
MAPPINGS = {
    "ISCC": {"Name": "Client Name", "Country": "Country", "Type": "Scope",
             "Certification Body": "Issuing CB", "Valid To": "Expiry Date"},
    "SURE": {"Name": "Client Name", "Country": "Country", "Type": "Scope",
             "Certification Body": "Issuing CB", "Valid To": "Expiry Date"},
    "PEFC": {"Identifier": ("Certificate Number", "Code"), "Name": "Entity",
             "Country": "Country", "Type": "Role", "Status": "Status"},
    "FSC": {"Identifier": "Certificate Code", "Name": "Organization",
            "Country": "Country", "Type": "Certificate Type", "Status": "Status",
            "Valid From": "Valid From", "Valid To": "Valid To"},
    "GGL": {"Identifier": "USI", "Name": "Participant name", "Country": "Country",
            "Type": "Participant role", "Certification Body": "CB",
            "Status": "Status", "Valid From": "Valid from", "Valid To": "Valid till"},
    "SBP": {"Identifier": "Certificate Number", "Name": "Certificate Holder",
            "Country": "Country", "Type": "Certificate Type",
            "Certification Body": "Certification Body", "Status": "Status",
            "Valid From": "Date of Issue", "Valid To": "Date of Expiry"},
}
DATE_FORMATS = ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d %B %Y", "%d %b %Y")

FULL_OUT = "All certificates latest.xlsx"
SLIM_OUT = "All certificates (dashboard) latest.xlsx"
COMBINED_CSV = "All certificates latest.csv"    # transient, feeds the Data sheet


def iso_date(v):
    """Normalise a scheme's date to an ISO 'YYYY-MM-DD' string (sorts uniformly);
    keep the raw value if it matches no known format."""
    if not v:
        return ""
    if isinstance(v, (datetime, date)):
        return (v.date() if isinstance(v, datetime) else v).isoformat()
    s = str(v).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s


def read_data(scheme):
    wb = load_workbook(f"{scheme} certificates latest.xlsx", read_only=True, data_only=True)
    try:
        ws = wb["Data"]
        it = ws.iter_rows(values_only=True)
        headers = list(next(it))
        rows = [["" if c is None else c for c in r]
                for r in it if r and not all(c is None or c == "" for c in r)]
        return headers, rows
    finally:
        wb.close()


def pick(row_map, source):
    if isinstance(source, tuple):
        return next((row_map[c] for c in source if row_map.get(c)), "")
    return row_map.get(source, "")


def add_detail_sheet(wb, scheme, headers, rows):
    ws = wb.create_sheet(scheme)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for c in range(1, len(headers) + 1):
        hdr(ws.cell(row=1, column=c))
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(max(len(str(h)) + 2, 12), 48)
    ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    tab = Table(displayName=f"{scheme}Data", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)
    ws.freeze_panes = "A2"


def add_summary_sheet(wb, per_scheme, total):
    ws = wb.create_sheet("Summary")
    ws.append(["Scheme", "Records"])
    for c in (1, 2):
        hdr(ws.cell(row=1, column=c))
    for scheme in SCHEMES:
        if scheme in per_scheme:
            ws.append([scheme, len(per_scheme[scheme][1])])
    ws.append(["TOTAL", total])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1).value = (
        "Certification Body is published only by ISCC, SURE, GGL and SBP "
        "(see the Certification Bodies sheet). PEFC and FSC do not publish it."
    )
    ws.cell(row=note_row, column=1).font = Font(italic=True, size=9, color="888888")


def add_cb_summary_sheet(wb, combined):
    """One row per certification body: record count + which schemes report it.
    PEFC/FSC carry no CB, so those rows are simply excluded (blank CB)."""
    from collections import defaultdict
    counts = defaultdict(int)
    schemes = defaultdict(set)
    for r in combined:
        cb = r[CB_I]
        cb = cb.strip() if isinstance(cb, str) else cb
        if not cb:
            continue
        counts[cb] += 1
        schemes[cb].add(r[SCHEME_I])

    ws = wb.create_sheet("Certification Bodies")
    ws.append(["Certification Body", "Records", "Schemes"])
    for c in (1, 2, 3):
        hdr(ws.cell(row=1, column=c))
    ordered = sorted(counts.items(), key=lambda x: (-x[1], x[0].lower()))
    for cb, cnt in ordered:
        ws.append([cb, cnt, ", ".join(sorted(schemes[cb]))])
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 24
    ws.freeze_panes = "A2"
    last = len(ordered) + 1
    tab = Table(displayName="CertBodies", ref=f"A1:C{last}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)
    return len(ordered)


def add_cb_to_dashboard(wb, combined, top_n=15):
    """Surface certification bodies on the Dashboard front page: a ranked table
    + horizontal bar of the top-N CBs by record count. Static (not tied to the
    Country/Scheme selectors) because ~89% of rows carry no CB, which would make
    an interactive per-country CB split mostly empty. Placed below the existing
    pie/bar charts so it doesn't overlap them."""
    from collections import defaultdict
    counts = defaultdict(int)
    for r in combined:
        cb = r[CB_I]
        cb = cb.strip() if isinstance(cb, str) else cb
        if cb:
            counts[cb] += 1
    top = sorted(counts.items(), key=lambda x: (-x[1], x[0].lower()))[:top_n]
    if not top:
        return

    ws = wb["Dashboard"]
    base = 50  # the pie (B19) and bar (E19) are 14 cm tall (~row 47); clear them

    ws.merge_cells(start_row=base, start_column=2, end_row=base, end_column=3)
    title = ws.cell(row=base, column=2)
    title.value = "Top Certification Bodies (ISCC · SURE · GGL · SBP publish a CB)"
    title.font = Font(bold=True, size=12, color=WHITE, name="Calibri")
    title.fill = PatternFill("solid", fgColor=BLUE)
    title.alignment = Alignment(horizontal="center", vertical="center")

    hr = base + 1
    hdr(ws.cell(row=hr, column=2)); ws.cell(row=hr, column=2).value = "Certification Body"
    hdr(ws.cell(row=hr, column=3)); ws.cell(row=hr, column=3).value = "Records"
    for i, (cb, cnt) in enumerate(top):
        rr = hr + 1 + i
        ws.cell(row=rr, column=2).value = cb
        ws.cell(row=rr, column=3).value = cnt
        bg = STRIPE if i % 2 == 0 else WHITE
        for c in (2, 3):
            ws.cell(row=rr, column=c).fill = PatternFill("solid", fgColor=bg)
            ws.cell(row=rr, column=c).font = Font(name="Calibri", size=10)

    first, last = hr + 1, hr + len(top)
    bar = BarChart()
    bar.type = "bar"            # horizontal, so the long CB names read on the axis
    bar.grouping = "clustered"
    bar.title = f"Top {len(top)} Certification Bodies by records"
    bar.style = 10
    bar.width = 22
    bar.height = 12
    bar.legend = None
    data = Reference(ws, min_col=3, min_row=hr, max_row=last)      # header for series title
    cats = Reference(ws, min_col=2, min_row=first, max_row=last)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    for s in bar.series:
        s.cat = AxDataSource(strRef=StrRef(f=f"'Dashboard'!$B${first}:$B${last}"))
    ws.add_chart(bar, f"E{base}")


def main():
    combined, per_scheme = [], {}
    for scheme in SCHEMES:
        if not glob.glob(f"{scheme} certificates latest.xlsx"):
            print(f"  skip {scheme}: dashboard not found")
            continue
        headers, rows = read_data(scheme)
        per_scheme[scheme] = (headers, rows)
        m = MAPPINGS[scheme]
        for r in rows:
            row_map = dict(zip(headers, r))
            rec = [scheme]
            for col in COMMON[1:]:
                if col not in m:
                    rec.append("")
                elif col in DATE_COLS:
                    rec.append(iso_date(pick(row_map, m[col])))
                else:
                    rec.append(pick(row_map, m[col]))
            combined.append(rec)
        print(f"  {scheme}: {len(rows)} rows")

    if not combined:
        sys.exit("No dashboards found — did the scrapers run?")

    # Feed the Data sheet + dashboard aggregation. dim2 ('cb' slot) = Scheme.
    with open(COMBINED_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(COMMON)
        w.writerows(combined)
    dash_rows = [{"country": r[COUNTRY_I] or "Unknown", "cb": r[SCHEME_I]} for r in combined]
    country_totals, cb_totals, cb_by_country, country_by_cb = aggregate(dash_rows)
    print(f"Building dashboard: {len(combined)} rows, {len(country_totals)} countries …")

    wb = build_excel(
        dash_rows, country_totals, cb_totals, cb_by_country, country_by_cb, COMBINED_CSV,
        title="All Certificates — Interactive Dashboard",
        dim2_singular="Scheme", dim2_short="Scheme",
        data_fieldnames=COMMON, data_widths=COMMON_WIDTHS,
        kpi_total_label="Total Certificate Records",
        default_prefix="All certificates", save=False,
    )
    add_cb_to_dashboard(wb, combined)
    n_cbs = add_cb_summary_sheet(wb, combined)
    add_summary_sheet(wb, per_scheme, len(combined))
    print(f"  Certification Bodies sheet: {n_cbs} bodies")

    # Slim (dashboard + data only) first — small enough to email.
    wb.save(SLIM_OUT)
    print(f"Saved → {SLIM_OUT}")

    # Then add per-scheme detail and save the full workbook (+ dated copy).
    for scheme in SCHEMES:
        if scheme in per_scheme:
            add_detail_sheet(wb, scheme, *per_scheme[scheme])
    # CI sets COMBINED_DATE once and reuses it for the commit step, so the
    # dated filename can't drift across a UTC-midnight boundary mid-run.
    date_str = os.environ.get("COMBINED_DATE") or datetime.now(timezone.utc).strftime("%Y.%m.%d")
    for path in (f"All certificates {date_str}.xlsx", FULL_OUT):
        wb.save(path)
        print(f"Saved → {path}")

    os.remove(COMBINED_CSV)


if __name__ == "__main__":
    main()
